#!/usr/bin/env python3
"""
Excelerator GL-bilag generator for UBW Agresso / Sigma2
Genererer xlsx-fil klar for import via Citrix/Excelerator.
"""

import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import os
import glob
import sys

DESKTOP = os.path.expanduser("~/Desktop")


def neste_batch_id(periode: int) -> int:
    """Finn høyeste brukte batch_id for perioden og inkrementer."""
    mønster = os.path.join(DESKTOP, f"*.xlsx")
    høyeste = -1
    for fil in glob.glob(mønster):
        if os.path.basename(fil).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(fil, data_only=True)
        except Exception:
            continue
        if "_control" in wb.sheetnames:
            ws = wb["_control"]
            for row in ws.iter_rows(values_only=True):
                if row[0] == "setdefault" and row[1] == "batch_id":
                    bid = row[2]
                    if isinstance(bid, int) and str(bid).startswith(str(periode)):
                        løpenr = int(str(bid)[6:])
                        høyeste = max(høyeste, løpenr)
    return int(f"{periode}{(høyeste + 1):02d}")


def lag_gl_bilag(
    periode: int,
    bokføringsdato: datetime,
    linjer: list[dict],
    bilagsnavn: str | None = None,
    batch_id: int | None = None,
) -> str:
    """
    Generer Excelerator GL-bilag som xlsx.

    Args:
        periode: YYYYMM (f.eks. 202603)
        bokføringsdato: datetime-objekt
        linjer: liste med dict per posteringslinje:
            {
              "account": int,        # kontonr (påkrevd)
              "dim_1": str | None,   # koststed
              "dim_2": str | None,   # prosjekt
              "dim_3": int | None,   # ressursnr / lev.nr
              "dim_4": str | None,
              "dim_6": str | None,
              "tax_code": str | None,
              "amount": float,       # positivt = debet, negativt = kredit
              "description": str,    # bilagstekst
            }
        bilagsnavn: filnavn uten .xlsx (valgfritt, autogenereres ellers)
        batch_id: overstyrer auto-beregnet batch_id

    Returns:
        Absolutt filsti til generert fil.
    """
    # Validering
    total = sum(l["amount"] for l in linjer)
    if abs(total) > 0.01:
        raise ValueError(f"Bilaget balanserer ikke: sum = {total:.2f} (må være 0)")
    if not linjer:
        raise ValueError("Ingen posteringslinjer")
    for i, l in enumerate(linjer):
        if "account" not in l or "amount" not in l or "description" not in l:
            raise ValueError(f"Linje {i+1} mangler påkrevde felt (account, amount, description)")

    if batch_id is None:
        batch_id = neste_batch_id(periode)

    wb = openpyxl.Workbook()

    # --- Ark 1: _control ---
    ws_ctrl = wb.active
    ws_ctrl.title = "_control"

    ws_ctrl["A1"] = "*"
    ws_ctrl["B1"] = "BOKFØRING AV BILAG FRA EXCEL"

    ws_ctrl["A4"] = "*"
    ws_ctrl["B4"] = "Global Parameters (setdefault will be used unless parameter of same name is passed in from Agresso)"

    ws_ctrl["A5"] = "*"
    ws_ctrl["B5"] = "Parameter"
    ws_ctrl["C5"] = "Value"

    params = [
        ("client",         "client",      "FIRMA KODE"),
        ("batch_id",       batch_id,      "BUNT NUMMER/ID"),
        ("period",         periode,       "Periode"),
        ("voucher_date",   bokføringsdato, "Bilagsdato"),
        ("voucher_no",     None,          "Bilagsnummer"),
        ("voucher_type",   "GL",          "Bilagsart"),
        ("user_id",        "user_id",     "Bruker ID"),
        ("currency",       "NOK",         "Valuta"),
        ("vouch_flag",     "Y",           "Skal GL07 tildele bilags nr? (Y/N)"),
        ("variant_number", 9,             "post back paramter for GL07 variant"),
        ("trans_type",     "GL",          "Hovedbilag = GL"),
        ("interface",      "BI",          "Forsystem"),
    ]
    for rad, (key, val, kommentar) in enumerate(params, start=6):
        ws_ctrl.cell(rad, 1, "setdefault")
        ws_ctrl.cell(rad, 2, key)
        ws_ctrl.cell(rad, 3, val)
        ws_ctrl.cell(rad, 4, kommentar)

    # --- Ark 2: Postering til UBW ---
    ws_post = wb.create_sheet("Postering til UBW")

    ws_post["C2"] = "Hovedbokstransaksjoner"

    kolonner = ["update_columns", "account", "dim_1", "dim_2", "dim_3",
                "dim_4", "dim_6", "tax_code", "amount", "cur_amount", "description"]
    for col, verdi in enumerate(kolonner, start=1):
        celle = ws_post.cell(9, col, verdi)
        celle.font = Font(bold=True)

    for rad_offset, linje in enumerate(linjer):
        rad = 10 + rad_offset
        ws_post.cell(rad, 1, "update_data")
        ws_post.cell(rad, 2, linje["account"])
        ws_post.cell(rad, 3, linje.get("dim_1"))
        ws_post.cell(rad, 4, linje.get("dim_2"))
        ws_post.cell(rad, 5, linje.get("dim_3"))
        ws_post.cell(rad, 6, linje.get("dim_4"))
        ws_post.cell(rad, 7, linje.get("dim_6"))
        ws_post.cell(rad, 8, linje.get("tax_code"))
        ws_post.cell(rad, 9, linje["amount"])
        ws_post.cell(rad, 10, linje["amount"])  # cur_amount = amount (NOK)
        ws_post.cell(rad, 11, linje["description"])

    # Filnavn
    if bilagsnavn is None:
        år = str(periode)[:4]
        mnd = str(periode)[4:6]
        bilagsnavn = f"{år} {mnd} Bilag {batch_id}"

    filsti = os.path.join(DESKTOP, f"{bilagsnavn}.xlsx")
    wb.save(filsti)
    return filsti


if __name__ == "__main__":
    # Eksempel: krysspostering av Basir-forskudd
    fil = lag_gl_bilag(
        periode=202603,
        bokføringsdato=datetime(2026, 3, 10),
        linjer=[
            {"account": 29300, "dim_1": None, "dim_2": None, "dim_3": None,
             "amount": -9789.0, "description": "Kryssing Stockholm 21.-23.10.25"},
            {"account": 29300, "dim_1": None, "dim_2": None, "dim_3": None,
             "amount": 9789.0,  "description": "Kryssing Stockholm 21.-23.10.25"},
        ],
        bilagsnavn="TEST Basir kryssing"
    )
    print(f"Generert: {fil}")
